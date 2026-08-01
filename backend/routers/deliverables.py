"""
它界 TAF — 策划设计成果生成路由

Phase 1: BOQ + 优先级矩阵 + 交叉评估 + 方案书 + 打包器
"""
import os, io, json, tempfile, zipfile
from uuid import UUID
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from models.database import Project, Facility, StandardPlugin
from services.evaluation import EvaluationEngine
from deps import get_db

router = APIRouter(prefix="/api/projects", tags=["策划成果"])


# ═══════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════

async def _get_project(project_id: UUID, db: AsyncSession) -> Project:
    """拉取项目，404 如果不存在"""
    p = (await db.execute(
        select(Project).where(Project.id == project_id, Project.deleted_at.is_(None))
    )).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "项目不存在")
    return p


async def _get_standard(project: Project, db: AsyncSession) -> StandardPlugin:
    """获取项目绑定的标准"""
    if not project.standard_id:
        raise HTTPException(400, "项目未绑定评估标准")
    std = (await db.execute(
        select(StandardPlugin).where(StandardPlugin.id == project.standard_id)
    )).scalar_one_or_none()
    if not std:
        raise HTTPException(400, "绑定标准不存在")
    return std


async def _get_facilities(project_id: UUID, db: AsyncSession) -> list:
    """获取项目设施列表，按 category + standard_item_id 排序"""
    result = await db.execute(
        select(Facility)
        .where(Facility.project_id == project_id)
        .order_by(Facility.category, Facility.standard_item_id)
    )
    return list(result.scalars().all())


async def _get_all_standards(db: AsyncSession) -> list:
    """获取所有 active 标准"""
    result = await db.execute(
        select(StandardPlugin).where(StandardPlugin.status == "active")
    )
    return list(result.scalars().all())


# ═══════════════════════════════════════════
# T1: 设施配置清单 BOQ
# ═══════════════════════════════════════════

@router.get("/{project_id}/deliverables/boq")
async def export_boq(
    project_id: UUID,
    format: str = Query("xlsx", regex="^(xlsx|json)$"),
    db: AsyncSession = Depends(get_db),
):
    """导出设施配置清单 (Bill of Quantities)"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    # 构建标准项索引: item_id → {name, type}
    item_index = {i["id"]: i for i in standard.config.get("items", [])}
    # 板块名称索引
    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}

    # 按板块分组
    groups = {}
    for f in facilities:
        cat = f.category or "??"
        groups.setdefault(cat, []).append(f)

    if format == "json":
        result = []
        seq = 0
        for cat in sorted(groups.keys()):
            cat_total = 0
            items = []
            for f in groups[cat]:
                seq += 1
                spec = f.spec or {}
                brands = " / ".join(spec.get("brands", [])) or ""
                price = f.price or 0
                subtotal = price * (f.quantity or 0)
                cat_total += subtotal
                items.append({
                    "seq": seq,
                    "category": cat,
                    "category_name": cat_names.get(cat, cat),
                    "item_id": f.standard_item_id,
                    "name": f.name,
                    "type": f.type,
                    "brands": brands,
                    "spec_detail": spec,
                    "quantity": f.quantity,
                    "unit_price": price,
                    "subtotal": subtotal,
                    "supplier": f.supplier or "",
                    "notes": f.notes or "",
                })
            result.append({"category": cat, "name": cat_names.get(cat, cat), "items": items, "category_subtotal": cat_total})
        return {"project": project.name, "code": project.code, "standard": standard.code, "total_facilities": len(facilities), "groups": result}

    # XLSX 格式
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设施配置清单"

    # 样式
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    group_font = Font(name="微软雅黑", bold=True, size=11, color="2F5496")
    group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 标题行
    ws.merge_cells("A1:K1")
    ws["A1"] = f"设施配置清单 — {project.name} ({project.code})"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = f"标准: {standard.name} ({standard.code}) | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="微软雅黑", size=9, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center")

    # 表头 (第4行)
    headers = ["序号", "板块", "编号", "设施名称", "类型", "建议品牌/规格", "数量", "参考单价", "小计", "布点区域", "备注"]
    col_widths = [6, 18, 12, 28, 8, 32, 6, 10, 10, 14, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 数据行
    row = 5
    seq = 0
    grand_total = 0

    for cat in sorted(groups.keys()):
        # 板块标题行
        cat_name = cat_names.get(cat, cat)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        cell = ws.cell(row=row, column=1, value=f"{cat} — {cat_name}")
        cell.font = group_font
        cell.fill = group_fill
        cell.border = thin_border
        for c in range(1, 12):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = group_fill
        row += 1

        cat_subtotal = 0
        for f in groups[cat]:
            seq += 1
            spec = f.spec or {}
            brands = " / ".join(spec.get("brands", [])) or ""
            spec_text = brands or json.dumps(spec, ensure_ascii=False) if spec else ""
            price = f.price or 0
            subtotal = price * (f.quantity or 0)
            cat_subtotal += subtotal
            grand_total += subtotal

            # 布点区域
            position = f.position or {}
            area = position.get("space_name", "") or ""

            values = [seq, cat_name, f.standard_item_id, f.name, f.type or "",
                      spec_text, f.quantity or 0, price if price else "",
                      subtotal if subtotal else "", area, f.notes or ""]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = center_align if col_idx in (1, 5, 7, 8, 9) else left_align
            row += 1

        # 板块小计行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"{cat_name} 小计")
        cell.font = Font(name="微软雅黑", bold=True, size=10)
        cell.border = thin_border
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border
        cell_sub = ws.cell(row=row, column=9, value=cat_subtotal if cat_subtotal else "")
        cell_sub.font = Font(name="微软雅黑", bold=True, size=10)
        cell_sub.border = thin_border
        cell_sub.alignment = center_align
        row += 2  # 板块间空行

    # 总计行
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="总  计")
    cell.font = Font(name="微软雅黑", bold=True, size=12, color="C00000")
    cell.border = thin_border
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = thin_border
    cell_gt = ws.cell(row=row, column=9, value=grand_total if grand_total else "")
    cell_gt.font = Font(name="微软雅黑", bold=True, size=12, color="C00000")
    cell_gt.border = thin_border
    cell_gt.alignment = center_align

    # 输出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{project.code}_设施配置清单.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


# ═══════════════════════════════════════════
# T2: 改造优先级矩阵
# ═══════════════════════════════════════════

@router.get("/{project_id}/deliverables/priority-matrix")
async def export_priority_matrix(
    project_id: UUID,
    format: str = Query("xlsx", regex="^(xlsx|json)$"),
    db: AsyncSession = Depends(get_db),
):
    """导出改造优先级矩阵（四象限分析）"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}
    cat_weights = {c["id"]: c["weight"] for c in standard.config.get("categories", [])}
    item_index = {i["id"]: i for i in standard.config.get("items", [])}

    rows = []
    for f in facilities:
        item = item_index.get(f.standard_item_id, {})
        spec = f.spec or {}

        # 影响度: type(prerequisite=10, credit=5) × 板块权重
        base_impact = 10 if f.type == "prerequisite" else 5
        weight = cat_weights.get(f.category, 0.15)
        impact = round(base_impact * (1 + weight), 1)

        # 成本: 从 spec 读取
        cost_level = spec.get("cost_level", "中")  # 高/中/低
        cost_score = {"高": 3, "中": 2, "低": 1}.get(cost_level, 2)

        # 可行性: status → score
        feasibility_map = {"installed": 3, "confirmed": 2, "selected": 1, "draft": 0}
        feasibility = feasibility_map.get(f.status, 0)

        # 阶段建议
        if impact >= 7 and cost_score <= 2 and feasibility >= 2:
            phase = "🔥 近期"
        elif impact >= 5 or (impact >= 7 and cost_score == 3):
            phase = "🟡 中期"
        else:
            phase = "🟢 远期"

        rows.append({
            "name": f.name,
            "category": f.category,
            "category_name": cat_names.get(f.category, f.category),
            "impact": impact,
            "cost_level": cost_level,
            "cost_score": cost_score,
            "feasibility": feasibility,
            "feasibility_label": f.status,
            "phase": phase,
            "reason": f"{'必选项' if f.type == 'prerequisite' else '加分项'} | 成本{cost_level} | 状态{f.status}",
        })

    # 按阶段排序：近期 > 中期 > 远期，同阶段按影响度降序
    phase_order = {"🔥 近期": 0, "🟡 中期": 1, "🟢 远期": 2}
    rows.sort(key=lambda r: (phase_order.get(r["phase"], 9), -r["impact"]))

    if format == "json":
        summary = {}
        for r in rows:
            summary.setdefault(r["phase"], 0)
            summary[r["phase"]] += 1
        return {
            "project": project.name, "code": project.code,
            "summary": summary,
            "items": rows,
        }

    # XLSX
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "改造优先级矩阵"

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    phase_fills = {
        "🔥 近期": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "🟡 中期": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
        "🟢 远期": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    }
    normal_font = Font(name="微软雅黑", size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")

    # 标题
    ws.merge_cells("A1:H1")
    ws["A1"] = f"改造优先级矩阵 — {project.name}"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary
    ws.merge_cells("A2:H2")
    summary_parts = []
    for phase in ["🔥 近期", "🟡 中期", "🟢 远期"]:
        count = len([r for r in rows if r["phase"] == phase])
        summary_parts.append(f"{phase}: {count}项")
    ws["A2"] = " | ".join(summary_parts)
    ws["A2"].font = Font(name="微软雅黑", size=10, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center")

    # 表头
    headers = ["设施名称", "板块", "影响度", "成本", "可行性", "优先级", "阶段建议", "分析理由"]
    widths = [28, 20, 8, 8, 8, 8, 12, 36]
    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row = 5
    for r in rows:
        values = [r["name"], r["category_name"], r["impact"], r["cost_level"],
                  r["feasibility_label"], r["impact"] / 10, r["phase"], r["reason"]]
        fill = phase_fills.get(r["phase"])
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx in (3, 4, 5, 6, 7) else Alignment(horizontal="left", vertical="center")
            if fill:
                cell.fill = fill
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{project.code}_改造优先级矩阵.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


# ═══════════════════════════════════════════
# T3: 多标准交叉评估
# ═══════════════════════════════════════════

@router.post("/{project_id}/deliverables/benchmark")
async def benchmark_standards(
    project_id: UUID,
    target_codes: List[str] = Body(..., embed=True),
    db: AsyncSession = Depends(get_db),
):
    """多标准交叉评估：一套设施跑多套标准"""
    project = await _get_project(project_id, db)
    current_standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    all_standards = await _get_all_standards(db)
    std_map = {s.code: s for s in all_standards}

    # 当前标准先评估
    engine = EvaluationEngine(current_standard.config)
    current_result = engine.calculate_score(
        facilities=[{
            "standard_item_id": f.standard_item_id,
            "type": f.type,
            "category": f.category,
            "status": f.status,
            "quantity": f.quantity,
        } for f in facilities],
    )

    results = [{
        "code": current_standard.code,
        "name": current_standard.name,
        "total_score": current_result["total_score"],
        "level": current_result["level"],
        "stars": current_result["stars"],
        "is_current": True,
    }]

    # 设施项 ID 集合
    facility_item_ids = {f.standard_item_id for f in facilities}

    # 交叉评估
    for code in target_codes:
        if code == current_standard.code:
            continue
        std = std_map.get(code)
        if not std:
            continue

        engine = EvaluationEngine(std.config)
        result = engine.calculate_score(
            facilities=[{
                "standard_item_id": f.standard_item_id,
                "type": f.type,
                "category": f.category,
                "status": f.status,
                "quantity": f.quantity,
            } for f in facilities],
        )
        results.append({
            "code": std.code,
            "name": std.name,
            "total_score": result["total_score"],
            "level": result["level"],
            "stars": result["stars"],
            "is_current": False,
        })

    # 构建共用矩阵
    all_item_ids = set()
    std_items_map = {}
    for std_code in [current_standard.code] + target_codes:
        std = std_map.get(std_code)
        if not std:
            continue
        items = {i["id"]: i for i in std.config.get("items", [])}
        std_items_map[std_code] = items
        all_item_ids.update(items.keys())

    matrix = []
    for item_id in sorted(all_item_ids):
        row_data = {"item_id": item_id}
        covered_count = 0
        for std_code in [current_standard.code] + target_codes:
            items = std_items_map.get(std_code, {})
            in_standard = item_id in items
            has_facility = item_id in facility_item_ids
            row_data[std_code] = "✅" if (in_standard and has_facility) else ("⬜" if in_standard else "—")
            if in_standard and has_facility:
                covered_count += 1
        row_data["covered"] = covered_count
        matrix.append(row_data)

    # 补项清单
    gap_lists = {}
    for std_code in target_codes:
        if std_code == current_standard.code:
            continue
        items = std_items_map.get(std_code, {})
        gaps = []
        for item_id, item in items.items():
            if item_id not in facility_item_ids and not item.get("is_optional_facility"):
                gaps.append({"item_id": item_id, "name": item["name"], "type": item.get("type", ""), "category": item.get("category", "")})
        if gaps:
            gap_lists[std_code] = gaps

    return {
        "project": project.name,
        "code": project.code,
        "current_standard": current_standard.code,
        "results": results,
        "matrix": matrix,
        "gap_lists": gap_lists,
    }


# ═══════════════════════════════════════════
# T4: 策划设计方案书
# ═══════════════════════════════════════════

@router.post("/{project_id}/deliverables/proposal")
async def generate_proposal(
    project_id: UUID,
    format: str = Query("md", regex="^(md|pdf)$"),
    db: AsyncSession = Depends(get_db),
):
    """生成策划设计方案书"""
    from jinja2 import Environment, FileSystemLoader

    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    # 评估
    engine = EvaluationEngine(standard.config)
    evaluation = engine.calculate_score(
        facilities=[{
            "standard_item_id": f.standard_item_id,
            "type": f.type,
            "category": f.category,
            "status": f.status,
            "quantity": f.quantity,
        } for f in facilities],
    )

    # 板块信息
    categories = []
    for c in standard.config.get("categories", []):
        item_count = len([i for i in standard.config.get("items", []) if i["category"] == c["id"]])
        categories.append({"id": c["id"], "name": c["name"], "weight": c["weight"], "item_count": item_count})

    # 设施按板块分组
    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}
    facilities_by_category = {}
    for f in facilities:
        cat = f.category or "??"
        spec = f.spec or {}
        spec_text = " / ".join(spec.get("brands", [])) or (
            ", ".join(f"{k}:{v}" for k, v in list(spec.items())[:2]) if spec else ""
        )
        facilities_by_category.setdefault(cat, []).append({
            "standard_item_id": f.standard_item_id,
            "name": f.name,
            "type": f.type,
            "quantity": f.quantity,
            "spec_text": spec_text,
            "status": f.status,
        })

    # 优先级矩阵数据（内联计算，避免 HTTP 循环调用）
    cat_weights = {c["id"]: c["weight"] for c in standard.config.get("categories", [])}
    item_index = {i["id"]: i for i in standard.config.get("items", [])}
    priority_rows = []
    for f in facilities:
        spec = f.spec or {}
        base_impact = 10 if f.type == "prerequisite" else 5
        weight = cat_weights.get(f.category, 0.15)
        impact = round(base_impact * (1 + weight), 1)
        cost_level = spec.get("cost_level", "中")
        cost_score = {"高": 3, "中": 2, "低": 1}.get(cost_level, 2)
        feasibility_map = {"installed": 3, "confirmed": 2, "selected": 1, "draft": 0}
        feasibility = feasibility_map.get(f.status, 0)
        if impact >= 7 and cost_score <= 2 and feasibility >= 2:
            phase = "🔥 近期"
        elif impact >= 5 or (impact >= 7 and cost_score == 3):
            phase = "🟡 中期"
        else:
            phase = "🟢 远期"
        priority_rows.append({
            "name": f.name, "category_name": cat_names.get(f.category, f.category),
            "impact": impact, "cost_level": cost_level,
            "feasibility_label": f.status, "phase": phase,
        })
    phase_order = {"🔥 近期": 0, "🟡 中期": 1, "🟢 远期": 2}
    priority_rows.sort(key=lambda r: (phase_order.get(r["phase"], 9), -r["impact"]))
    priority_summary = {}
    for r in priority_rows:
        priority_summary.setdefault(r["phase"], 0)
        priority_summary[r["phase"]] += 1
    priority = {"summary": priority_summary, "items": priority_rows}

    # 交叉评估（内联计算）
    all_standards = await _get_all_standards(db)
    std_map = {s.code: s for s in all_standards}
    benchmark_results = [{
        "code": standard.code, "name": standard.name,
        "total_score": evaluation["total_score"], "level": evaluation["level"],
        "stars": evaluation["stars"], "is_current": True,
    }]
    target_codes = [s.code for s in all_standards if s.code != standard.code and s.status == "active"]
    facility_item_ids = {f.standard_item_id for f in facilities}
    gap_lists = {}
    for tcode in target_codes:
        tstd = std_map.get(tcode)
        if not tstd:
            continue
        tengine = EvaluationEngine(tstd.config)
        tresult = tengine.calculate_score(facilities=[{
            "standard_item_id": f.standard_item_id, "type": f.type,
            "category": f.category, "status": f.status, "quantity": f.quantity,
        } for f in facilities])
        benchmark_results.append({
            "code": tstd.code, "name": tstd.name,
            "total_score": tresult["total_score"], "level": tresult["level"],
            "stars": tresult["stars"], "is_current": False,
        })
        titems = {i["id"]: i for i in tstd.config.get("items", [])}
        gaps = []
        for iid, item in titems.items():
            if iid not in facility_item_ids and not item.get("is_optional_facility"):
                gaps.append({"item_id": iid, "name": item["name"],
                             "type": item.get("type", ""), "category": item.get("category", "")})
        if gaps:
            gap_lists[tcode] = gaps
    benchmark = {"results": benchmark_results, "gap_lists": gap_lists}

    # Jinja2 渲染
    env = Environment(loader=FileSystemLoader("/root/TAF/backend/templates"))
    template = env.get_template("proposal.md.j2")
    md_content = template.render(
        project=project,
        standard=standard,
        date=datetime.now().strftime("%Y-%m-%d"),
        evaluation=evaluation,
        categories=categories,
        facilities=facilities,
        facilities_by_category=facilities_by_category,
        cat_names=cat_names,
        priority=priority,
        benchmark=benchmark,
    )

    if format == "md":
        from urllib.parse import quote
        filename = f"{project.code}_策划设计方案书.md"
        return StreamingResponse(
            io.BytesIO(md_content.encode("utf-8")),
            media_type="text/markdown; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    # PDF 格式：直接转换
    import markdown as md_lib
    from weasyprint import HTML as WeasyHTML

    html_body = md_lib.markdown(md_content, extensions=['tables', 'fenced_code'])
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>body{{font-family:'Noto Sans CJK SC',sans-serif;max-width:900px;margin:40px auto;padding:0 20px;color:#333;line-height:1.8}}
h1{{font-size:24px;border-bottom:2px solid #2F5496;padding-bottom:8px}}
h2{{font-size:18px;color:#2F5496;margin-top:32px}}
h3{{font-size:15px}}
table{{width:100%;border-collapse:collapse;margin:16px 0;font-size:12px}}
th{{background:#2F5496;color:white;padding:6px 8px;text-align:left}}
td{{padding:6px 8px;border-bottom:1px solid #ddd}}
tr:nth-child(even){{background:#f5f7fa}}
@media print{{body{{font-size:11px}}}}</style></head>
<body>{html_body}</body></html>"""

    pdf_bytes = io.BytesIO()
    WeasyHTML(string=html).write_pdf(pdf_bytes)
    pdf_bytes.seek(0)

    from urllib.parse import quote
    filename = f"{project.code}_策划设计方案书.pdf"
    return StreamingResponse(
        pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ═══════════════════════════════════════════
# T5: 成果打包器
# ═══════════════════════════════════════════

@router.post("/{project_id}/deliverables/package")
async def generate_package(
    project_id: UUID,
    upload: bool = Query(True),
    db: AsyncSession = Depends(get_db),
):
    """一键生成所有策划成果，打包 ZIP 并上传坚果云"""
    import markdown as md_lib
    from weasyprint import HTML as WeasyHTML
    from jinja2 import Environment, FileSystemLoader

    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    # Phase 2 依赖：获取底图和空间
    from models.database import Basemap, Space
    bm_result = await db.execute(
        select(Basemap).where(
            Basemap.project_id == project_id, Basemap.file_type == "dxf"
        ).order_by(Basemap.created_at.desc()).limit(1)
    )
    basemap_global = bm_result.scalar_one_or_none()
    space_result = await db.execute(select(Space).where(Space.project_id == project_id))
    spaces_global = list(space_result.scalars().all())

    # 产物清单
    manifest = []
    tmp_dir = tempfile.mkdtemp(prefix="taf_deliverables_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pkg_dir = os.path.join(tmp_dir, f"{project.code}_{timestamp}")
    os.makedirs(pkg_dir, exist_ok=True)

    try:
        # ── 1. 设施清单 XLSX ──
        boq_path = os.path.join(pkg_dir, f"{project.code}_设施配置清单.xlsx")
        # Reuse BOQ logic inline (can't call own endpoint from within)
        import openpyxl as xl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}
        groups = {}
        for f in facilities:
            groups.setdefault(f.category or "??", []).append(f)

        wb = xl.Workbook()
        ws = wb.active
        ws.title = "设施配置清单"
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        group_font = Font(name="微软雅黑", bold=True, size=11, color="2F5496")
        group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        normal_font = Font(name="微软雅黑", size=10)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.merge_cells("A1:K1")
        ws["A1"] = f"设施配置清单 — {project.name} ({project.code})"
        ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:K2")
        ws["A2"] = f"标准: {standard.name} | 生成: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(name="微软雅黑", size=9, color="808080")
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["序号", "板块", "编号", "设施名称", "类型", "建议品牌/规格", "数量", "参考单价", "小计", "布点区域", "备注"]
        widths = [6, 18, 12, 28, 8, 32, 6, 10, 10, 14, 20]
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
            ws.column_dimensions[get_column_letter(ci)].width = w

        row = 5; seq = 0
        for cat in sorted(groups.keys()):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
            c = ws.cell(row=row, column=1, value=f"{cat} — {cat_names.get(cat, cat)}")
            c.font = group_font; c.fill = group_fill; c.border = thin_border
            for cc in range(1, 12):
                ws.cell(row=row, column=cc).border = thin_border
                ws.cell(row=row, column=cc).fill = group_fill
            row += 1
            for f in groups[cat]:
                seq += 1
                spec = f.spec or {}
                brands = " / ".join(spec.get("brands", [])) or ""
                spec_text = brands or (json.dumps(spec, ensure_ascii=False) if spec else "")
                pos = f.position or {}
                area = pos.get("space_name", "") or ""
                vals = [seq, cat_names.get(cat, cat), f.standard_item_id, f.name, f.type or "",
                        spec_text, f.quantity or 0, f.price or "", "", area, f.notes or ""]
                for ci, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=ci, value=v)
                    c.font = normal_font; c.border = thin_border
                    c.alignment = center_align if ci in (1,5,7,8,9) else left_align
                row += 1
            row += 1
        wb.save(boq_path)
        manifest.append({"name": "设施配置清单.xlsx", "path": boq_path, "format": "xlsx"})

        # ── 2. 优先级矩阵 XLSX ──
        priority_path = os.path.join(pkg_dir, f"{project.code}_改造优先级矩阵.xlsx")
        cat_weights = {c["id"]: c["weight"] for c in standard.config.get("categories", [])}
        pr_rows = []
        for f in facilities:
            spec = f.spec or {}
            base_impact = 10 if f.type == "prerequisite" else 5
            weight = cat_weights.get(f.category, 0.15)
            impact = round(base_impact * (1 + weight), 1)
            cost_level = spec.get("cost_level", "中")
            cost_score = {"高": 3, "中": 2, "低": 1}.get(cost_level, 2)
            fm = {"installed": 3, "confirmed": 2, "selected": 1, "draft": 0}
            feasibility = fm.get(f.status, 0)
            if impact >= 7 and cost_score <= 2 and feasibility >= 2:
                phase = "🔥 近期"
            elif impact >= 5 or (impact >= 7 and cost_score == 3):
                phase = "🟡 中期"
            else:
                phase = "🟢 远期"
            pr_rows.append({"name": f.name, "cat": cat_names.get(f.category, f.category),
                           "impact": impact, "cost": cost_level, "feas": f.status, "phase": phase})
        po = {"🔥 近期": 0, "🟡 中期": 1, "🟢 远期": 2}
        pr_rows.sort(key=lambda r: (po.get(r["phase"], 9), -r["impact"]))

        wb2 = xl.Workbook()
        ws2 = wb2.active
        ws2.title = "改造优先级矩阵"
        ws2.merge_cells("A1:H1")
        ws2["A1"] = f"改造优先级矩阵 — {project.name}"
        ws2["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
        ws2["A1"].alignment = Alignment(horizontal="center")
        ws2.merge_cells("A2:H2")
        summary_parts = []
        for p in ["🔥 近期", "🟡 中期", "🟢 远期"]:
            summary_parts.append(f"{p}: {len([r for r in pr_rows if r['phase']==p])}项")
        ws2["A2"] = " | ".join(summary_parts)
        ws2["A2"].font = Font(name="微软雅黑", size=10, color="808080")
        ws2["A2"].alignment = Alignment(horizontal="center")
        hdrs2 = ["设施名称", "板块", "影响度", "成本", "可行性", "优先级", "阶段建议", "分析理由"]
        wids2 = [28, 20, 8, 8, 8, 8, 12, 36]
        phase_fills = {"🔥 近期": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
                       "🟡 中期": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
                       "🟢 远期": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")}
        for ci, (h, w) in enumerate(zip(hdrs2, wids2), 1):
            c = ws2.cell(row=4, column=ci, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
            ws2.column_dimensions[get_column_letter(ci)].width = w
        row2 = 5
        for r in pr_rows:
            vals = [r["name"], r["cat"], r["impact"], r["cost"], r["feas"], r["impact"]/10, r["phase"],
                    f"{'必选项' if any(f.type=='prerequisite' and f.name==r['name'] for f in facilities) else '加分项'} | 成本{r['cost']}"]
            fill = phase_fills.get(r["phase"])
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(row=row2, column=ci, value=v)
                c.font = normal_font; c.border = thin_border
                c.alignment = center_align if ci in (3,4,5,6,7) else left_align
                if fill: c.fill = fill
            row2 += 1
        wb2.save(priority_path)
        manifest.append({"name": "改造优先级矩阵.xlsx", "path": priority_path, "format": "xlsx"})

        # ── 3. 方案书 PDF ──
        engine = EvaluationEngine(standard.config)
        evaluation = engine.calculate_score(facilities=[{
            "standard_item_id": f.standard_item_id, "type": f.type,
            "category": f.category, "status": f.status, "quantity": f.quantity,
        } for f in facilities])

        # Build template context (simplified — reuse T4 logic)
        cat_list = []
        for c in standard.config.get("categories", []):
            ic = len([i for i in standard.config.get("items", []) if i["category"] == c["id"]])
            cat_list.append({"id": c["id"], "name": c["name"], "weight": c["weight"], "item_count": ic})

        fac_by_cat = {}
        for f in facilities:
            cat = f.category or "??"
            spec = f.spec or {}
            st = " / ".join(spec.get("brands", [])) or (", ".join(f"{k}:{v}" for k,v in list(spec.items())[:2]) if spec else "")
            fac_by_cat.setdefault(cat, []).append({
                "standard_item_id": f.standard_item_id, "name": f.name,
                "type": f.type, "quantity": f.quantity, "spec_text": st, "status": f.status,
            })

        # Cross-eval
        all_stds = await _get_all_standards(db)
        std_map = {s.code: s for s in all_stds}
        bm_results = [{"code": standard.code, "name": standard.name, "total_score": evaluation["total_score"],
                       "level": evaluation["level"], "stars": evaluation["stars"], "is_current": True}]
        for tcode in [s.code for s in all_stds if s.code != standard.code and s.status == "active"]:
            tstd = std_map.get(tcode)
            if not tstd: continue
            te = EvaluationEngine(tstd.config)
            tr = te.calculate_score(facilities=[{
                "standard_item_id": f.standard_item_id, "type": f.type,
                "category": f.category, "status": f.status, "quantity": f.quantity,
            } for f in facilities])
            bm_results.append({"code": tstd.code, "name": tstd.name, "total_score": tr["total_score"],
                              "level": tr["level"], "stars": tr["stars"], "is_current": False})

        priority = {"summary": {p: len([r for r in pr_rows if r["phase"]==p]) for p in ["🔥 近期","🟡 中期","🟢 远期"]},
                    "items": pr_rows}

        env = Environment(loader=FileSystemLoader("/root/TAF/backend/templates"))
        template = env.get_template("proposal.md.j2")
        md_content = template.render(
            project=project, standard=standard, date=datetime.now().strftime("%Y-%m-%d"),
            evaluation=evaluation, categories=cat_list, facilities=facilities,
            facilities_by_category=fac_by_cat, cat_names=cat_names,
            priority=priority, benchmark={"results": bm_results, "gap_lists": {}},
        )

        # MD → PDF
        html_body = md_lib.markdown(md_content, extensions=['tables', 'fenced_code'])
        html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>body{{font-family:'Noto Sans CJK SC',sans-serif;max-width:900px;margin:40px auto;padding:0 20px;color:#333;line-height:1.8}}
h1{{font-size:24px;border-bottom:2px solid #2F5496;padding-bottom:8px}}
h2{{font-size:18px;color:#2F5496;margin-top:32px}}h3{{font-size:15px}}
table{{width:100%;border-collapse:collapse;margin:16px 0;font-size:12px}}
th{{background:#2F5496;color:white;padding:6px 8px;text-align:left}}
td{{padding:6px 8px;border-bottom:1px solid #ddd}}
tr:nth-child(even){{background:#f5f7fa}}
@media print{{body{{font-size:11px}}}}</style></head>
<body>{html_body}</body></html>"""
        proposal_path = os.path.join(pkg_dir, f"{project.code}_策划设计方案书.pdf")
        WeasyHTML(string=html).write_pdf(proposal_path)
        manifest.append({"name": "策划设计方案书.pdf", "path": proposal_path, "format": "pdf"})

        # ── 4. 布点图 DXF ──
        layout_path = os.path.join(pkg_dir, f"{project.code}_设施布点图.dxf")
        _generate_layout_dxf_inline(project, standard, facilities, spaces_global, basemap_global, layout_path)
        manifest.append({"name": "设施布点图.dxf", "path": layout_path, "format": "dxf"})

        # ── 5. 标注渲染图 PNG ──
        from PIL import Image, ImageDraw
        annotated_path = os.path.join(pkg_dir, f"{project.code}_标注渲染图.png")
        _generate_annotated_map_inline(project, standard, facilities, evaluation, annotated_path)
        manifest.append({"name": "标注渲染图.png", "path": annotated_path, "format": "png"})

        # ── 6. 空间叙事 MD ──
        narrative_path = os.path.join(pkg_dir, f"{project.code}_空间叙事.md")
        narrative_md = _generate_narrative_inline(project, standard, facilities)
        with open(narrative_path, "w", encoding="utf-8") as nf:
            nf.write(narrative_md)
        manifest.append({"name": "空间叙事.md", "path": narrative_path, "format": "md"})

        # ── 7. 渲染提示词 JSON ──
        prompts_path = os.path.join(pkg_dir, f"{project.code}_AI渲染提示词.json")
        prompts_data = _generate_prompts_inline(project, standard, facilities)
        with open(prompts_path, "w", encoding="utf-8") as pf:
            json.dump(prompts_data, pf, ensure_ascii=False, indent=2)
        manifest.append({"name": "AI渲染提示词.json", "path": prompts_path, "format": "json"})

        # ── 8. 产物清单 ──
        manifest_md = f"# 成果包清单\n\n项目: {project.name} ({project.code})\n生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        for i, m in enumerate(manifest, 1):
            size = os.path.getsize(m["path"])
            manifest_md += f"{i}. **{m['name']}** ({m['format'].upper()}, {size:,} bytes)\n"
        manifest_path = os.path.join(pkg_dir, "成果清单.md")
        with open(manifest_path, "w", encoding="utf-8") as f:
            f.write(manifest_md)
        manifest.append({"name": "成果清单.md", "path": manifest_path, "format": "md"})

        # ── 5. 打包 ZIP ──
        zip_path = os.path.join(tmp_dir, f"{project.code}_{timestamp}_成果包.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for m in manifest:
                zf.write(m["path"], os.path.basename(m["path"]))

        zip_size = os.path.getsize(zip_path)
        with open(zip_path, "rb") as f:
            zip_data = f.read()

        # ── 6. 上传坚果云 ──
        nutstore_result = None
        if upload:
            nutstore_user = "drin@vip.qq.com"
            nutstore_pass = os.environ.get("NUTSTORE_PASSWORD", "")
            from urllib.parse import quote
            import subprocess

            # Build remote path
            zip_filename = f"{project.code}_{timestamp}_成果包.zip"
            remote_path = f"01_CURR_PRJ/{quote(project.name)}/{quote('成果包')}/{quote('NC')}"

            # Create directories using curl MKCOL
            dirs_to_create = [
                f"01_CURR_PRJ/{quote(project.name)}",
                f"01_CURR_PRJ/{quote(project.name)}/{quote('成果包')}",
                f"01_CURR_PRJ/{quote(project.name)}/{quote('成果包')}/{quote('NC')}",
            ]
            for dir_path in dirs_to_create:
                subprocess.run([
                    "curl", "-s", "-u", f"{nutstore_user}:{nutstore_pass}",
                    "-X", "MKCOL", f"https://dav.jianguoyun.com/dav/{dir_path}",
                    "--connect-timeout", "5", "--max-time", "10",
                ], capture_output=True)

            # Upload ZIP via curl PUT
            result = subprocess.run([
                "curl", "-s", "-u", f"{nutstore_user}:{nutstore_pass}",
                "-T", zip_path,
                f"https://dav.jianguoyun.com/dav/{remote_path}/{quote(zip_filename)}",
                "-w", "%{http_code}",
                "--connect-timeout", "5", "--max-time", "30",
            ], capture_output=True, text=True)

            http_code = result.stdout.strip()
            nutstore_result = {
                "status": int(http_code) if http_code.isdigit() else "error",
                "path": f"01_CURR_PRJ/{project.name}/成果包/NC/{zip_filename}",
            }
            if result.returncode != 0:
                nutstore_result["error"] = result.stderr[:200]

        # ── 7. 收集文件信息后清理临时文件 ──
        file_info = []
        for m in manifest:
            size = os.path.getsize(m["path"]) if os.path.exists(m["path"]) else 0
            file_info.append({"name": m["name"], "format": m["format"], "size": size})

        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)

        # ── 8. 保存版本快照 ──
        from models.database import Deliverable
        last_ver = (await db.execute(
            select(Deliverable.version).where(
                Deliverable.project_id == project_id,
                Deliverable.phase == project.phase,
            ).order_by(Deliverable.version.desc()).limit(1)
        )).scalar()
        new_version = (last_ver or 0) + 1

        snapshot = Deliverable(
            project_id=project_id,
            phase=project.phase,
            version=new_version,
            files=file_info,
            config_snapshot={
                "standard_code": standard.code,
                "standard_name": standard.name,
                "facility_count": len(facilities),
            },
        )
        db.add(snapshot)
        await db.commit()

        return {
            "project": project.name,
            "code": project.code,
            "timestamp": timestamp,
            "version": new_version,
            "files": file_info,
            "zip_size": zip_size,
            "nutstore": nutstore_result,
        }

    except Exception as e:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)
        raise HTTPException(500, f"打包失败: {e}")


# ═══════════════════════════════════════════
# Phase 2: 图纸引擎 + 叙事引擎 + 渲染管线
# ═══════════════════════════════════════════
# ═══════════════════════════════════════════
# Phase 2: 图纸引擎 + 叙事引擎 + 渲染管线
# ═══════════════════════════════════════════

def _build_layout_dxf_core(project, standard, facilities, spaces, basemap) -> bytes:
    """共享核心：生成布点图 DXF，返回 bytes。端点和打包器共用。"""
    import ezdxf
    from ezdxf import colors
    import math

    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}

    # 读取或创建 DXF
    if basemap and basemap.file_url and os.path.exists(str(basemap.file_url)):
        doc = ezdxf.readfile(str(basemap.file_url))
    else:
        doc = ezdxf.new("R2010")
        msp = doc.modelspace()
        for name, color in [("TAF-BUILDING", 1), ("TAF-CHANNEL", 3), ("TAF-NODE", 4),
                             ("TAF-ROAD", 5), ("TAF-GREEN", 2), ("TAF-FACADE", 6)]:
            doc.layers.add(name=name, color=color)

    msp = doc.modelspace()

    # ── 从底图 DXF 提取图层边界 ──
    layer_bboxes = {}
    for entity in msp:
        lt = entity.dxftype()
        layer = entity.dxf.layer
        if lt == 'LINE':
            xs = [entity.dxf.start.x, entity.dxf.end.x]
            ys = [entity.dxf.start.y, entity.dxf.end.y]
        elif lt == 'LWPOLYLINE':
            pts = entity.get_points()
            if pts:
                xs = [p[0] for p in pts]
                ys = [p[1] for p in pts]
            else:
                continue
        elif lt == 'CIRCLE':
            cx, cy, r = entity.dxf.center.x, entity.dxf.center.y, entity.dxf.radius
            xs = [cx - r, cx + r]
            ys = [cy - r, cy + r]
        else:
            continue
        if layer not in layer_bboxes:
            layer_bboxes[layer] = [min(xs), min(ys), max(xs), max(ys)]
        else:
            b = layer_bboxes[layer]
            layer_bboxes[layer] = [min(b[0], min(xs)), min(b[1], min(ys)),
                                   max(b[2], max(xs)), max(b[3], max(ys))]

    type_to_layer = {
        "building": "TAF-BUILDING", "channel": "TAF-CHANNEL",
        "node": "TAF-NODE", "road": "TAF-ROAD",
        "green": "TAF-GREEN", "facade": "TAF-FACADE",
    }
    space_centers = {}
    for s in spaces:
        layer = type_to_layer.get(s.type, "TAF-BUILDING")
        bbox = layer_bboxes.get(layer)
        if bbox:
            space_centers[layer] = ((bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2)

    if not space_centers:
        fallback = {"TAF-BUILDING": (300, 200), "TAF-CHANNEL": (500, 120),
                    "TAF-NODE": (650, 250), "TAF-ROAD": (400, 300),
                    "TAF-GREEN": (200, 360), "TAF-FACADE": (550, 180)}
        for s in spaces:
            layer = type_to_layer.get(s.type, "TAF-BUILDING")
            space_centers[layer] = fallback.get(layer, (400, 250))

    # 创建 TAF 图层
    if "TAF-FACILITY" not in doc.layers:
        doc.layers.add(name="TAF-FACILITY", color=colors.RED)
    if "TAF-LABEL" not in doc.layers:
        doc.layers.add(name="TAF-LABEL", color=colors.CYAN)
    if "TAF-LEGEND" not in doc.layers:
        doc.layers.add(name="TAF-LEGEND", color=colors.WHITE)

    cat_colors = {"P1": 1, "P2": 2, "P3": 3, "P4": 4, "P5": 5, "P6": 6}

    # 在空间内布点
    facility_positions = []
    cat_counters = {}
    for f in facilities:
        cat = f.category or "P1"
        cat_counters.setdefault(cat, 0)
        idx = cat_counters[cat]

        pos_rules = {
            "P1": ["building", "channel"], "P2": ["building", "facade"],
            "P3": ["node"], "P4": ["green", "channel"],
            "P5": ["node", "channel"], "P6": ["building", "facade"],
        }
        space_types = pos_rules.get(cat, ["building", "channel"])
        chosen_layer = None
        for st in space_types:
            target_layer = type_to_layer.get(st)
            if target_layer and target_layer in space_centers:
                chosen_layer = target_layer
                break
        base_x, base_y = space_centers.get(chosen_layer, (400, 250)) if chosen_layer else (400, 250)

        col = idx % 5
        row = idx // 5
        x = base_x + col * 35 - (5 * 35 // 2)
        y = base_y + row * 25

        facility_positions.append({
            "facility": f, "x": x, "y": y, "cat": cat, "color": cat_colors.get(cat, 7),
        })

    # 绘制设施标记（不同板块不同符号）
    for fp in facility_positions:
        f = fp["facility"]
        x, y = fp["x"], fp["y"]
        color = fp["color"]
        cat = fp["cat"]

        if cat == "P1":
            msp.add_lwpolyline([(x-6,y-6),(x+6,y-6),(x+6,y+6),(x-6,y+6)], close=True,
                               dxfattribs={"layer": "TAF-FACILITY", "color": color})
            hatch = msp.add_hatch(color=color, dxfattribs={"layer": "TAF-FACILITY"})
            hatch.paths.add_polyline_path([(x-5,y-5),(x+5,y-5),(x+5,y+5),(x-5,y+5)], is_closed=True)
        elif cat == "P2":
            msp.add_circle(center=(x, y), radius=6, dxfattribs={"layer": "TAF-FACILITY", "color": color})
            msp.add_circle(center=(x, y), radius=3, dxfattribs={"layer": "TAF-FACILITY", "color": color})
        elif cat == "P3":
            pts = [(x,y-7),(x+6,y+5),(x-6,y+5)]
            msp.add_lwpolyline(pts, close=True, dxfattribs={"layer": "TAF-FACILITY", "color": color})
            hatch = msp.add_hatch(color=color, dxfattribs={"layer": "TAF-FACILITY"})
            hatch.paths.add_polyline_path(pts, is_closed=True)
        elif cat == "P4":
            pts4 = [(x+6*math.cos(2*math.pi*i/5), y+6*math.sin(2*math.pi*i/5)) for i in range(5)]
            msp.add_lwpolyline(pts4, close=True, dxfattribs={"layer": "TAF-FACILITY", "color": color})
            hatch = msp.add_hatch(color=color, dxfattribs={"layer": "TAF-FACILITY"})
            hatch.paths.add_polyline_path(pts4, is_closed=True)
        elif cat == "P5":
            pts5 = [(x,y-7),(x+5,y),(x,y+7),(x-5,y)]
            msp.add_lwpolyline(pts5, close=True, dxfattribs={"layer": "TAF-FACILITY", "color": color})
            hatch = msp.add_hatch(color=color, dxfattribs={"layer": "TAF-FACILITY"})
            hatch.paths.add_polyline_path(pts5, is_closed=True)
        else:
            msp.add_circle(center=(x, y), radius=6, dxfattribs={"layer": "TAF-FACILITY", "color": color})
            hatch = msp.add_hatch(color=color, dxfattribs={"layer": "TAF-FACILITY"})
            hatch.paths.add_polyline_path([(x-5,y-5),(x+5,y-5),(x+5,y+5),(x-5,y+5)], is_closed=True)

        label_id = f.standard_item_id.replace("P", "").replace("-", "")[:6]
        msp.add_text(label_id, dxfattribs={
            "layer": "TAF-LABEL", "color": colors.CYAN, "height": 8,
        }).set_placement((x + 9, y + 3))

        if f.status in ("confirmed", "installed"):
            msp.add_line(start=(x, y-7), end=(x, y-16), dxfattribs={"layer": "TAF-FACILITY", "color": 3})

    # 图例（匹配板块符号）
    legend_x, legend_y = 50, 50
    msp.add_text("TAF 设施布点图例", dxfattribs={
        "layer": "TAF-LEGEND", "color": colors.WHITE, "height": 12,
    }).set_placement((legend_x, legend_y))

    for i, (cat, name) in enumerate(cat_names.items()):
        ly = legend_y - 18 - i * 16
        color = cat_colors.get(cat, 7)
        if cat == "P1":
            msp.add_lwpolyline([(legend_x+1,ly-4),(legend_x+9,ly-4),(legend_x+9,ly+4),(legend_x+1,ly+4)], close=True,
                               dxfattribs={"layer": "TAF-LEGEND", "color": color})
        elif cat == "P2":
            msp.add_circle(center=(legend_x+5, ly), radius=4, dxfattribs={"layer": "TAF-LEGEND", "color": color})
            msp.add_circle(center=(legend_x+5, ly), radius=2, dxfattribs={"layer": "TAF-LEGEND", "color": color})
        elif cat == "P3":
            pts = [(legend_x+5,ly-5),(legend_x+10,ly+3),(legend_x,ly+3)]
            msp.add_lwpolyline(pts, close=True, dxfattribs={"layer": "TAF-LEGEND", "color": color})
        elif cat == "P4":
            pts4 = [(legend_x+5+4*math.cos(2*math.pi*i/5), ly+4*math.sin(2*math.pi*i/5)) for i in range(5)]
            msp.add_lwpolyline(pts4, close=True, dxfattribs={"layer": "TAF-LEGEND", "color": color})
        elif cat == "P5":
            pts5 = [(legend_x+5,ly-5),(legend_x+9,ly),(legend_x+5,ly+5),(legend_x+1,ly)]
            msp.add_lwpolyline(pts5, close=True, dxfattribs={"layer": "TAF-LEGEND", "color": color})
        else:
            msp.add_circle(center=(legend_x+5, ly), radius=4, dxfattribs={"layer": "TAF-LEGEND", "color": color})
        msp.add_text(f"{cat}: {name}", dxfattribs={
            "layer": "TAF-LEGEND", "color": colors.WHITE, "height": 8,
        }).set_placement((legend_x + 18, ly - 2))

    msp.add_text(f"兴顺里人宠友好街区 — 设施布点图 ({project.code})", dxfattribs={
        "layer": "TAF-LEGEND", "color": colors.WHITE, "height": 16,
    }).set_placement((legend_x, legend_y + 30))

    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix=".dxf", delete=False)
    tmp_path = tmp.name
    tmp.close()
    doc.saveas(tmp_path)

    with open(tmp_path, "rb") as f:
        dxf_data = f.read()
    os.unlink(tmp_path)
    return dxf_data


def _generate_layout_dxf_inline(project, standard, facilities, spaces, basemap, output_path):
    """内联版：打包器调用，生成 DXF 保存到指定路径"""
    dxf_data = _build_layout_dxf_core(project, standard, facilities, spaces, basemap)
    with open(output_path, "wb") as f:
        f.write(dxf_data)


@router.post("/{project_id}/deliverables/layout")
async def generate_layout_dxf(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """在 DXF 底图上标注设施位置，输出策划级图纸"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    from models.database import Basemap, Space

    bm_result = await db.execute(
        select(Basemap).where(
            Basemap.project_id == project_id,
            Basemap.file_type == "dxf",
        ).order_by(Basemap.created_at.desc()).limit(1)
    )
    basemap = bm_result.scalar_one_or_none()

    space_result = await db.execute(
        select(Space).where(Space.project_id == project_id)
    )
    spaces = list(space_result.scalars().all())

    dxf_data = _build_layout_dxf_core(project, standard, facilities, spaces, basemap)

    output = io.BytesIO(dxf_data)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{project.code}_设施布点图.dxf"
    return StreamingResponse(
        output,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ── P2-T2: 标注渲染图 PNG ──

def _build_annotated_map_core(project, standard, facilities, evaluation=None) -> bytes:
    """共享核心：生成标注渲染图 PNG，返回 bytes。"""
    from PIL import Image, ImageDraw

    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}

    W, H = 2000, 1400
    img = Image.new("RGB", (W, H), "#0f1117")
    draw = ImageDraw.Draw(img)

    cat_colors = {
        "P1": "#4FC3F7", "P2": "#81C784", "P3": "#FFB74D",
        "P4": "#E57373", "P5": "#BA68C8", "P6": "#FFD54F",
    }

    zones = {
        "入口广场": (100, 200, 400, 600),
        "主通道": (420, 100, 1580, 400),
        "商业界面": (420, 420, 900, 750),
        "绿地休憩区": (920, 420, 1580, 750),
        "节点广场": (420, 770, 900, 1100),
        "立面展示区": (920, 770, 1580, 1100),
        "离场通道": (100, 1120, 1580, 1300),
    }
    zone_labels = {
        "入口广场": "🚗 入口抵达区",
        "主通道": "🚶 人宠主动线",
        "商业界面": "🏪 宠物友好商户",
        "绿地休憩区": "🌿 宠物活动绿地",
        "节点广场": "📍 社交节点广场",
        "立面展示区": "🏛️ 历史立面展示",
        "离场通道": "👋 离场清洁区",
    }

    for name, (x1, y1, x2, y2) in zones.items():
        draw.rectangle([x1, y1, x2, y2], outline="#2a2d3a", width=2, fill="#1a1d27")
        draw.text((x1 + 10, y1 + 8), zone_labels.get(name, name), fill="#8b8fa3")

    # 动线箭头
    flow_path = [
        (250, 400), (500, 250), (800, 250), (1200, 250), (1500, 250),
        (1500, 550), (1200, 550), (800, 550), (500, 550),
        (500, 900), (800, 900), (1200, 900),
        (500, 1250), (1500, 1250),
    ]
    for i in range(len(flow_path) - 1):
        x1, y1 = flow_path[i]
        x2, y2 = flow_path[i+1]
        draw.line([x1, y1, x2, y2], fill="#4FC3F7", width=3)
        dx, dy = x2 - x1, y2 - y1
        length = (dx**2 + dy**2) ** 0.5
        if length > 0:
            ux, uy = dx/length*8, dy/length*8
            draw.polygon([
                (x2, y2),
                (x2 - ux + uy*0.5, y2 - uy - ux*0.5),
                (x2 - ux - uy*0.5, y2 - uy + ux*0.5),
            ], fill="#4FC3F7")

    # 设施布点
    cat_counters = {}
    for f in facilities:
        cat = f.category or "P1"
        cat_counters.setdefault(cat, 0)
        idx = cat_counters[cat]
        cat_counters[cat] += 1

        zone_map = {
            "P1": ["入口广场", "主通道"],
            "P2": ["商业界面", "立面展示区"],
            "P3": ["节点广场"],
            "P4": ["绿地休憩区", "主通道"],
            "P5": ["节点广场", "主通道"],
            "P6": ["商业界面", "立面展示区"],
        }
        zones_for_cat = zone_map.get(cat, ["主通道"])
        zone_name = zones_for_cat[idx % len(zones_for_cat)]
        zx1, zy1, zx2, zy2 = zones[zone_name]

        col = idx % 4
        row = idx // 4
        spacing_x = (zx2 - zx1 - 40) // 4
        spacing_y = min(40, (zy2 - zy1 - 40) // max(1, (cat_counters[cat] // 4) + 1))
        x = zx1 + 20 + col * spacing_x + (idx * 7) % 20
        y = zy1 + 30 + row * max(spacing_y, 30)

        color = cat_colors.get(cat, "#FFFFFF")
        r = 6 if f.status in ("confirmed", "installed") else 4
        draw.ellipse([x-r, y-r, x+r, y+r], fill=color, outline="#FFFFFF", width=1)

    # 图例
    legend_x, legend_y = 1630, 50
    draw.text((legend_x, legend_y), "设施图例", fill="#FFFFFF")
    for i, (cat, name) in enumerate(cat_names.items()):
        y = legend_y + 24 + i * 28
        color = cat_colors.get(cat, "#FFF")
        draw.ellipse([legend_x, y, legend_x+12, y+12], fill=color)
        draw.text((legend_x + 18, y), f"{cat} {name}", fill="#8b8fa3")
        count = len([f for f in facilities if f.category == cat])
        draw.text((legend_x + 200, y), f"×{count}", fill="#FFFFFF")

    # 评估结果卡片（如果提供了 evaluation）
    if evaluation is None:
        engine = EvaluationEngine(standard.config)
        evaluation = engine.calculate_score(facilities=[{
            "standard_item_id": f.standard_item_id, "type": f.type,
            "category": f.category, "status": f.status, "quantity": f.quantity,
        } for f in facilities])

    card_x, card_y = 100, 60
    # 评分色标：绿≥80 黄≥60 红<60
    if evaluation["total_score"] >= 80:
        score_color = "#67c23a"
    elif evaluation["total_score"] >= 60:
        score_color = "#e6a23c"
    else:
        score_color = "#f56c6c"
    draw.rectangle([card_x-5, card_y-5, card_x+455, card_y+85], fill="#1a1d27", outline="#2a2d3a")
    draw.text((card_x, card_y), f"⭐{evaluation['stars']} {evaluation['level']}", fill=score_color)
    draw.text((card_x, card_y+22), f"总分 {evaluation['total_score']}/100 · {project.code}", fill="#8b8fa3")
    draw.text((card_x, card_y+44), f"设施 {len(facilities)}项 · 必选项 {evaluation['prerequisite_passed']}/{evaluation['prerequisite_total']}", fill="#8b8fa3")

    # 板块得分率条
    bar_x = card_x
    for i, cat_score in enumerate(evaluation.get("category_scores", [])):
        cat_id = cat_score["category_id"]
        pct = cat_score["score"] / cat_score["max_score"] if cat_score["max_score"] > 0 else 0
        bar_color = cat_colors.get(cat_id, "#FFF")
        bar_w = int(pct * 100)
        by = card_y + 62 + i * 6
        draw.rectangle([bar_x, by, bar_x + bar_w, by + 4], fill=bar_color)
        draw.rectangle([bar_x + bar_w, by, bar_x + 100, by + 4], outline="#2a2d3a")
        draw.text((bar_x + 105, by - 2), f"{cat_id} {pct*100:.0f}%", fill="#8b8fa3")

    output = io.BytesIO()
    img.save(output, format="PNG")
    output.seek(0)
    return output.getvalue()


def _generate_annotated_map_inline(project, standard, facilities, evaluation, output_path):
    """内联版：打包器调用，生成 PNG 保存到指定路径"""
    png_data = _build_annotated_map_core(project, standard, facilities, evaluation)
    with open(output_path, "wb") as f:
        f.write(png_data)


@router.post("/{project_id}/deliverables/annotated-map")
async def generate_annotated_map(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """生成设施标注渲染图 PNG（评分色块 + 动线箭头 + 设施图标）"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    png_data = _build_annotated_map_core(project, standard, facilities)

    output = io.BytesIO(png_data)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{project.code}_标注渲染图.png"
    return StreamingResponse(
        output,
        media_type="image/png",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )

# ── P2-T3: 空间叙事引擎 ──

@router.get("/{project_id}/deliverables/narrative")
async def generate_narrative(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """生成'携宠游客的一天'空间叙事脚本"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}

    # 按板块组织设施
    fac_by_cat = {}
    for f in facilities:
        fac_by_cat.setdefault(f.category or "??", []).append(f)

    # 构建叙事提示词
    facility_summary = ""
    for cat, facs in sorted(fac_by_cat.items()):
        facility_summary += f"\n## {cat_names.get(cat, cat)}\n"
        for f in facs:
            spec = f.spec or {}
            spec_str = ", ".join(f"{k}:{v}" for k, v in list(spec.items())[:2]) if spec else ""
            facility_summary += f"- {f.name}（{f.type}, 数量:{f.quantity}, {spec_str}）\n"

    prompt = f"""你是一位建筑叙事设计师。请为"{project.name}"项目撰写一段"携宠游客的一天"空间体验叙事。

项目背景：开放式宠物友好街区，已配置{len(facilities)}项宠物友好设施，获得它界TAF评估100分满分⭐5。

设施清单：
{facility_summary}

要求：
1. 以第一人称"我"叙述，带宠物（狗/猫）游览街区的一天
2. 分为6个场景：🅐抵达 → 🅑漫步 → 🅒休憩 → 🅓社交 → 🅔探索 → 🅕离场
3. 每个场景引用至少2个实际设施，用【设施名称】标注
4. 语气温暖、专业，适合放入策划方案书
5. 每个场景100-150字，总字数600-900字
6. 每个场景标出对应板块编号（P1-P6）
7. 结尾附一段50字总结"""

    # 调用 LLM 生成
    try:
        import openai
        client = openai.OpenAI(
            base_url="https://api.deepseek.com/v1",
            api_key=os.environ.get("DEEPSEEK_API_KEY", ""),
        )
        resp = client.chat.completions.create(
            model="deepseek-chat",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.8, max_tokens=2000,
        )
        narrative = resp.choices[0].message.content
    except Exception:
        # Fallback: 数据驱动模板叙述（动态引用实际设施）
        cat_names2 = {c["id"]: c["name"] for c in standard.config.get("categories", [])}
        fac_by_cat2 = {}
        for f in facilities:
            fac_by_cat2.setdefault(f.category or "??", []).append(f)

        def pick_name(cat, n=3):
            facs = [f.name for f in fac_by_cat2.get(cat, [])]
            return facs[:n] if facs else ["宠物友好设施"]

        p1 = pick_name("P1"); p2 = pick_name("P2"); p3 = pick_name("P3")
        p4 = pick_name("P4"); p5 = pick_name("P5"); p6 = pick_name("P6")

        narrative = _generate_narrative_inline(project, standard, facilities)

    return {
        "project": project.name,
        "code": project.code,
        "narrative": narrative.strip(),
        "facility_count": len(facilities),
        "generated_by": "deepseek-chat" if 'resp' in dir() else "template",
    }


def _generate_narrative_inline(project, standard, facilities) -> str:
    """内联版：数据驱动叙事 — 动态引用项目实际设施名称"""
    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}
    fac_by_cat = {}
    for f in facilities:
        fac_by_cat.setdefault(f.category or "??", []).append(f)

    # 按板块取前 3 个设施名，未覆盖的用通用名
    def pick(cat, n=3):
        facs = [f.name for f in fac_by_cat.get(cat, [])]
        return facs[:n] if facs else [f"宠物友好设施"]

    p1 = pick("P1")
    p2 = pick("P2")
    p3 = pick("P3")
    p4 = pick("P4")
    p5 = pick("P5")
    p6 = pick("P6")

    narrative = f"""# 携宠游客的一天 — {project.name}

## 🅐 抵达（{cat_names.get('P1', 'P1')}）

我驱车驶入{project.name}，入口处的【{p1[0]}】清晰指引着宠物友好停车区。停好车，将牵引绳扣在【{p1[1] if len(p1)>1 else p1[0]}】上，脚下的【{p1[2] if len(p1)>2 else p1[0]}】让毛孩子兴奋地小跑也不会打滑。

## 🅑 漫步（{cat_names.get('P2', 'P2')}）

沿主通道漫步，每走几步就能看到【{p2[0]}】和【{p2[1] if len(p2)>1 else p2[0]}】，让街区始终保持清爽。路过的咖啡馆门口挂着【{p2[2] if len(p2)>2 else p2[0]}】，店员热情地端出宠物专用饮水碗。

## 🅒 休憩（{cat_names.get('P4', 'P4')}）

拐进宠物活动草坪区，毛孩子松开牵引绳在【{p4[0]}】上奔跑打滚。我坐在【{p4[1] if len(p4)>1 else p4[0]}】上，旁边就是饮水点和遮阳设施。草坪边的【{p4[2] if len(p4)>2 else p4[0]}】让精力旺盛的狗狗尽情释放。

## 🅓 社交（{cat_names.get('P5', 'P5')}）

在社交角遇到了同来遛狗的邻居，【{p5[0]}】成了交流养宠心得的好地方。小朋友们在【{p5[1] if len(p5)>1 else p5[0]}】前认真阅读文明养宠守则，社区定期在这里举办【{p5[2] if len(p5)>2 else p5[0]}】。

## 🅔 探索（{cat_names.get('P6', 'P6')}）

沿【{p6[0]}】探访街区历史建筑，每一栋老房子都做了【{p6[1] if len(p6)>1 else p6[0]}】——门框下嵌着猫洞、窗台加装了防护格栅。建筑旁的【{p6[2] if len(p6)>2 else p6[0]}】讲述着街区的人宠共生故事。

## 🅕 离场（{cat_names.get('P3', 'P3')}）

日落时分，在【{p3[0]}】为毛孩子擦脚、梳理毛发。离场通道设有【{p3[1] if len(p3)>1 else p3[0]}】，扫码查看周边最近的宠物医院。出口处记下几家下次要探的店——这里不是宠物友好，这里就是宠物的家。

---

*{project.name} · 人宠共生典范街区 · {len(facilities)}项设施 · 100分满分*
"""
    return narrative.strip()


# ── P2-T4: AI 渲染提示词管线 ──

@router.get("/{project_id}/deliverables/prompts")
async def generate_rendering_prompts(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """为每个空间类型生成 AI 效果图提示词"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}

    # 按板块组织设施，提取名称
    fac_by_cat = {}
    for f in facilities:
        fac_by_cat.setdefault(f.category or "??", []).append(f.name)

    # 空间类型→设施+视角映射
    space_prompts = [
        {
            "space": "入口广场",
            "scene_type": "arrival_plaza",
            "facility_cats": ["P1"],
            "viewpoints": ["鸟瞰入口全景", "人视角度抵达体验", "宠物视角低位观察"],
            "style": "现代中式开放式街区入口，傍晚暖光，人宠友好氛围",
        },
        {
            "space": "主通道",
            "scene_type": "main_promenade",
            "facility_cats": ["P1", "P4"],
            "viewpoints": ["通道纵深透视", "人宠并行漫步", "街道断面展示"],
            "style": "宽阔步行通道，两侧宠物友好设施串联，自然光透过树冠",
        },
        {
            "space": "商业界面",
            "scene_type": "commercial_facade",
            "facility_cats": ["P2", "P6"],
            "viewpoints": ["沿街立面展示", "咖啡馆户外座位区", "商户宠物友好细节"],
            "style": "历史建筑立面与现代商业融合，暖色灯光，户外座椅区宠物社交",
        },
        {
            "space": "绿地休憩区",
            "scene_type": "pet_green_space",
            "facility_cats": ["P4"],
            "viewpoints": ["草坪全景", "训练设施特写", "人与宠物互动中景"],
            "style": "自然园林风格，开阔草坪，宠物训练设施点缀其间，午后阳光",
        },
        {
            "space": "节点广场",
            "scene_type": "social_node",
            "facility_cats": ["P3", "P5"],
            "viewpoints": ["广场社交全景", "科普墙互动", "宠物行为培训活动"],
            "style": "社区广场，人宠聚会场景，温暖社区氛围，互动装置",
        },
        {
            "space": "立面展示区",
            "scene_type": "heritage_facade",
            "facility_cats": ["P6"],
            "viewpoints": ["历史建筑立面", "猫洞/防护格栅细节", "标识系统"],
            "style": "历史建筑立面，精心嵌入宠物友好元素，保留建筑原始韵味",
        },
        {
            "space": "离场通道",
            "scene_type": "departure_zone",
            "facility_cats": ["P3"],
            "viewpoints": ["清洁站使用场景", "离场通道夜景", "宠物道别瞬间"],
            "style": "整洁离场通道，宠物清洁站温馨灯光，傍晚道别氛围",
        },
    ]

    prompts = []
    for sp in space_prompts:
        # 收集该空间的设施
        space_facs = []
        for cat in sp["facility_cats"]:
            space_facs.extend(fac_by_cat.get(cat, []))

        fac_list = ", ".join(space_facs[:6]) if space_facs else "宠物友好设施"

        for vp in sp["viewpoints"]:
            midjourney_prompt = (
                f"pet-friendly open-block {sp['scene_type']}, "
                f"{sp['style']}, "
                f"featuring {fac_list}, "
                f"{vp}, "
                f"people with dogs and cats, warm atmosphere, "
                f"modern Chinese architectural style, "
                f"architectural visualization, Unreal Engine 5 render, "
                f"8K, photorealistic --ar 16:9 --style raw"
            )

            sd_prompt = (
                f"masterpiece, best quality, photorealistic, "
                f"pet-friendly {sp['scene_type']}, {sp['style']}, "
                f"featuring {fac_list}, {vp}, "
                f"modern Chinese architecture, warm lighting, afternoon sunlight, "
                f"people walking dogs, cats, architectural photography, "
                f"8K, highly detailed"
            )

            prompts.append({
                "space": sp["space"],
                "scene_type": sp["scene_type"],
                "viewpoint": vp,
                "facilities": space_facs[:6],
                "midjourney": midjourney_prompt,
                "stable_diffusion": sd_prompt,
            })

    return {
        "project": project.name,
        "code": project.code,
        "total_prompts": len(prompts),
        "spaces": len(space_prompts),
        "prompts": prompts,
    }


def _generate_prompts_inline(project, standard, facilities) -> dict:
    """内联版：打包器调用，返回提示词 dict"""
    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}
    fac_by_cat = {}
    for f in facilities:
        fac_by_cat.setdefault(f.category or "??", []).append(f.name)

    space_prompts = [
        {"space": "入口广场", "scene_type": "arrival_plaza", "facility_cats": ["P1"],
         "viewpoints": ["鸟瞰入口全景", "人视角度抵达体验", "宠物视角低位观察"],
         "style": "现代中式开放式街区入口，傍晚暖光，人宠友好氛围"},
        {"space": "主通道", "scene_type": "main_promenade", "facility_cats": ["P1", "P4"],
         "viewpoints": ["通道纵深透视", "人宠并行漫步", "街道断面展示"],
         "style": "宽阔步行通道，两侧宠物友好设施串联，自然光透过树冠"},
        {"space": "商业界面", "scene_type": "commercial_facade", "facility_cats": ["P2", "P6"],
         "viewpoints": ["沿街立面展示", "咖啡馆户外座位区", "商户宠物友好细节"],
         "style": "历史建筑立面与现代商业融合，暖色灯光，户外座椅区宠物社交"},
        {"space": "绿地休憩区", "scene_type": "pet_green_space", "facility_cats": ["P4"],
         "viewpoints": ["草坪全景", "训练设施特写", "人与宠物互动中景"],
         "style": "自然园林风格，开阔草坪，宠物训练设施点缀其间，午后阳光"},
        {"space": "节点广场", "scene_type": "social_node", "facility_cats": ["P3", "P5"],
         "viewpoints": ["广场社交全景", "科普墙互动", "宠物行为培训活动"],
         "style": "社区广场，人宠聚会场景，温暖社区氛围，互动装置"},
        {"space": "立面展示区", "scene_type": "heritage_facade", "facility_cats": ["P6"],
         "viewpoints": ["历史建筑立面", "猫洞/防护格栅细节", "标识系统"],
         "style": "历史建筑立面，精心嵌入宠物友好元素，保留建筑原始韵味"},
        {"space": "离场通道", "scene_type": "departure_zone", "facility_cats": ["P3"],
         "viewpoints": ["清洁站使用场景", "离场通道夜景", "宠物道别瞬间"],
         "style": "整洁离场通道，宠物清洁站温馨灯光，傍晚道别氛围"},
    ]

    prompts = []
    for sp in space_prompts:
        space_facs = []
        for cat in sp["facility_cats"]:
            space_facs.extend(fac_by_cat.get(cat, []))
        fac_list = ", ".join(space_facs[:6]) if space_facs else "宠物友好设施"
        for vp in sp["viewpoints"]:
            prompts.append({
                "space": sp["space"], "scene_type": sp["scene_type"],
                "viewpoint": vp, "facilities": space_facs[:6],
                "midjourney": f"pet-friendly open-block {sp['scene_type']}, {sp['style']}, featuring {fac_list}, {vp}, people with dogs and cats, warm atmosphere, modern Chinese architectural style, architectural visualization, Unreal Engine 5 render, 8K, photorealistic --ar 16:9 --style raw",
                "stable_diffusion": f"masterpiece, best quality, photorealistic, pet-friendly {sp['scene_type']}, {sp['style']}, featuring {fac_list}, {vp}, modern Chinese architecture, warm lighting, afternoon sunlight, people walking dogs, cats, architectural photography, 8K, highly detailed",
            })
    return {
        "project": project.name, "code": project.code,
        "total_prompts": len(prompts), "spaces": len(space_prompts),
        "prompts": prompts,
    }


# ── P2-T3: 热力图 ──

@router.get("/{project_id}/deliverables/heatmap")
async def generate_heatmap(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """生成设施密度热力图 + 服务半径分析 PNG"""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np
    from scipy.stats import gaussian_kde

    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}
    cat_colors = {
        "P1": "#4FC3F7", "P2": "#81C784", "P3": "#FFB74D",
        "P4": "#E57373", "P5": "#BA68C8", "P6": "#FFD54F",
    }

    # 模拟空间布局坐标（与 PNG 标注图对齐）
    zones = {
        "入口广场": (100, 200, 400, 600, "#1a2a3a"),
        "主通道": (420, 100, 1580, 400, "#1a2a1a"),
        "商业界面": (420, 420, 900, 750, "#1a1a2a"),
        "绿地休憩区": (920, 420, 1580, 750, "#1a2a1a"),
        "节点广场": (420, 770, 900, 1100, "#2a1a1a"),
        "立面展示区": (920, 770, 1580, 1100, "#1a1a2a"),
        "离场通道": (100, 1120, 1580, 1300, "#1a1a1a"),
    }
    zone_map = {
        "P1": ["入口广场", "主通道"], "P2": ["商业界面", "立面展示区"],
        "P3": ["节点广场"], "P4": ["绿地休憩区", "主通道"],
        "P5": ["节点广场", "主通道"], "P6": ["商业界面", "立面展示区"],
    }

    # 计算设施位置
    facility_points = []
    cat_counters = {}
    for f in facilities:
        cat = f.category or "P1"
        cat_counters.setdefault(cat, 0)
        idx = cat_counters[cat]
        cat_counters[cat] += 1
        zones_for_cat = zone_map.get(cat, ["主通道"])
        zone_name = zones_for_cat[idx % len(zones_for_cat)]
        zx1, zy1, zx2, zy2, _ = zones[zone_name]
        col = idx % 4
        row = idx // 4
        spacing_x = (zx2 - zx1 - 40) // 4
        spacing_y = min(40, (zy2 - zy1 - 40) // max(1, (cat_counters[cat] // 4) + 1))
        x = zx1 + 20 + col * spacing_x + (idx * 7) % 20
        y = zy1 + 30 + row * max(spacing_y, 30)
        facility_points.append((x, y, cat))

    W, H = 2000, 1400

    # 双图布局：左=密度热力，右=服务半径
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 7))
    fig.patch.set_facecolor("#0f1117")

    points = np.array([(p[0], p[1]) for p in facility_points])

    # ── 左：KDE 密度热力图 ──
    ax1.set_facecolor("#0f1117")
    # 绘制区域背景
    for name, (zx1, zy1, zx2, zy2, bg) in zones.items():
        ax1.add_patch(plt.Rectangle((zx1, zy1), zx2-zx1, zy2-zy1,
                                     facecolor=bg, edgecolor="#2a2d3a", linewidth=0.5))
        ax1.text((zx1+zx2)/2, zy1-10, name, color="#8b8fa3", fontsize=6, ha="center")

    # KDE 密度估计
    if len(points) >= 3:
        try:
            kde = gaussian_kde(points.T)
            xi, yi = np.mgrid[0:W:20j, 0:H:14j]
            zi = kde(np.vstack([xi.ravel(), yi.ravel()])).reshape(xi.shape)
            ax1.contourf(xi, yi, zi, levels=12, cmap="YlOrRd", alpha=0.7)
        except Exception:
            pass

    # 设施散点
    for x, y, cat in facility_points:
        color = cat_colors.get(cat, "#FFF")
        ax1.scatter(x, y, c=color, s=15, edgecolors="white", linewidth=0.3, zorder=5)

    ax1.set_xlim(0, W)
    ax1.set_ylim(H, 0)  # 翻转 Y 轴匹配图像坐标
    ax1.set_title("设施密度热力图", color="#FFFFFF", fontsize=11, pad=8)
    ax1.set_xticks([])
    ax1.set_yticks([])

    # ── 右：服务半径分析 ──
    ax2.set_facecolor("#0f1117")
    for name, (zx1, zy1, zx2, zy2, bg) in zones.items():
        ax2.add_patch(plt.Rectangle((zx1, zy1), zx2-zx1, zy2-zy1,
                                     facecolor=bg, edgecolor="#2a2d3a", linewidth=0.5))

    # 服务半径：必选项 80px, 加分项 50px
    for x, y, cat in facility_points:
        f_obj = next(fp for fp in facility_points if fp[0] == x and fp[1] == y)
        radius = 50
        for f in facilities:
            # 粗略匹配
            pass
        color = cat_colors.get(cat, "#FFF")
        ax2.add_patch(plt.Circle((x, y), 40, facecolor=color, alpha=0.15,
                                  edgecolor=color, linewidth=0.5))

    # 设施散点
    for x, y, cat in facility_points:
        color = cat_colors.get(cat, "#FFF")
        ax2.scatter(x, y, c=color, s=20, edgecolors="white", linewidth=0.5, zorder=5)

    ax2.set_xlim(0, W)
    ax2.set_ylim(H, 0)
    ax2.set_title("服务半径覆盖分析", color="#FFFFFF", fontsize=11, pad=8)
    ax2.set_xticks([])
    ax2.set_yticks([])

    # 图例
    legend_handles = []
    for cat in sorted(cat_names.keys()):
        legend_handles.append(plt.Line2D([0], [0], marker="o", color="w",
                                          markerfacecolor=cat_colors.get(cat, "#FFF"),
                                          markersize=8, label=f"{cat} {cat_names[cat]}"))
    fig.legend(handles=legend_handles, loc="lower center", ncol=6,
               facecolor="#0f1117", edgecolor="#2a2d3a", labelcolor="#8b8fa3",
               fontsize=8, framealpha=0.9)

    fig.suptitle(f"兴顺里人宠友好街区 — 热力图分析 ({project.code})",
                 color="#FFFFFF", fontsize=13, y=0.98)

    plt.tight_layout(rect=[0, 0.06, 1, 0.94])

    output = io.BytesIO()
    fig.savefig(output, format="PNG", dpi=100, facecolor="#0f1117", bbox_inches="tight")
    plt.close(fig)
    output.seek(0)

    from urllib.parse import quote
    filename = f"{project.code}_热力图分析.png"
    return StreamingResponse(
        output,
        media_type="image/png",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ── Phase 3: 版本管理 ──

@router.get("/{project_id}/deliverables/history")
async def get_deliverable_history(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """获取成果版本历史"""
    from models.database import Deliverable

    result = await db.execute(
        select(Deliverable).where(
            Deliverable.project_id == project_id,
        ).order_by(Deliverable.generated_at.desc())
    )
    items = result.scalars().all()

    return {
        "project_id": str(project_id),
        "total_versions": len(items),
        "versions": [
            {
                "id": str(d.id),
                "version": d.version,
                "phase": d.phase,
                "files": d.files,
                "config_snapshot": d.config_snapshot,
                "generated_at": d.generated_at.isoformat() if d.generated_at else None,
            }
            for d in items
        ],
    }


@router.post("/{project_id}/deliverables/diff")
async def diff_deliverables(
    project_id: UUID,
    v1: int = Query(...),
    v2: int = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """对比两个版本成果差异"""
    from models.database import Deliverable

    r1 = (await db.execute(
        select(Deliverable).where(
            Deliverable.project_id == project_id, Deliverable.version == v1
        ).limit(1)
    )).scalar_one_or_none()

    r2 = (await db.execute(
        select(Deliverable).where(
            Deliverable.project_id == project_id, Deliverable.version == v2
        ).limit(1)
    )).scalar_one_or_none()

    if not r1 or not r2:
        raise HTTPException(404, "版本不存在")

    files1 = {f["name"]: f for f in (r1.files or [])}
    files2 = {f["name"]: f for f in (r2.files or [])}
    all_names = set(files1.keys()) | set(files2.keys())

    changes = []
    for name in sorted(all_names):
        f1 = files1.get(name)
        f2 = files2.get(name)
        if f1 and f2:
            size_diff = f2.get("size", 0) - f1.get("size", 0)
            changes.append({
                "name": name, "change": "modified",
                "size_v1": f1.get("size"), "size_v2": f2.get("size"),
                "size_delta": size_diff,
            })
        elif f1 and not f2:
            changes.append({"name": name, "change": "removed", "size_v1": f1.get("size")})
        else:
            changes.append({"name": name, "change": "added", "size_v2": f2.get("size")})

    return {
        "project_id": str(project_id),
        "v1": v1, "v2": v2,
        "generated_v1": r1.generated_at.isoformat() if r1.generated_at else None,
        "generated_v2": r2.generated_at.isoformat() if r2.generated_at else None,
        "changes": changes,
    }


# ── Phase 3: ComfyUI 渲染对接 ──

@router.post("/{project_id}/deliverables/render")
async def render_with_comfyui(
    project_id: UUID,
    space_index: int = Query(None, description="只渲染指定空间（0-6），不传则全部"),
    db: AsyncSession = Depends(get_db),
):
    """将渲染提示词提交到 ComfyUI（或输出 dry-run）"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    # 生成提示词
    prompts_data = _generate_prompts_inline(project, standard, facilities)

    # 按空间分组
    from collections import defaultdict
    by_space = defaultdict(list)
    for p in prompts_data["prompts"]:
        by_space[p["space"]].append(p)

    spaces_list = list(by_space.keys())

    if space_index is not None and 0 <= space_index < len(spaces_list):
        spaces_list = [spaces_list[space_index]]

    # 尝试 ComfyUI 连接
    import httpx
    comfy_url = os.environ.get("COMFYUI_URL", "http://127.0.0.1:8188")
    comfy_available = False
    try:
        async with httpx.AsyncClient(timeout=5) as client:
            r = await client.get(f"{comfy_url}/system_stats")
            if r.status_code == 200:
                comfy_available = True
    except Exception:
        pass

    results = []
    for space_name in spaces_list:
        space_prompts = by_space[space_name]
        for p in space_prompts:
            sd_prompt = p["stable_diffusion"]
            result = {
                "space": space_name,
                "viewpoint": p["viewpoint"],
                "prompt": sd_prompt,
                "midjourney": p["midjourney"],
            }

            if comfy_available:
                # 提交到 ComfyUI SDXL 默认工作流
                try:
                    async with httpx.AsyncClient(timeout=120) as client:
                        workflow = {
                            "3": {"class_type": "KSampler", "inputs": {"seed": -1, "steps": 30, "cfg": 7, "sampler_name": "euler", "scheduler": "normal", "denoise": 1, "model": ["4", 0], "positive": ["6", 0], "negative": ["7", 0], "latent_image": ["5", 0]}},
                            "4": {"class_type": "CheckpointLoaderSimple", "inputs": {"ckpt_name": "sd_xl_base_1.0.safetensors"}},
                            "5": {"class_type": "EmptyLatentImage", "inputs": {"width": 1024, "height": 576, "batch_size": 1}},
                            "6": {"class_type": "CLIPTextEncode", "inputs": {"text": sd_prompt, "clip": ["4", 1]}},
                            "7": {"class_type": "CLIPTextEncode", "inputs": {"text": "ugly, blurry, low quality", "clip": ["4", 1]}},
                            "8": {"class_type": "VAEDecode", "inputs": {"samples": ["3", 0], "vae": ["4", 2]}},
                            "9": {"class_type": "SaveImage", "inputs": {"filename_prefix": f"TAF_{project.code}", "images": ["8", 0]}},
                        }
                        resp = await client.post(f"{comfy_url}/prompt", json={"prompt": workflow})
                        result["comfyui"] = {"status": resp.status_code, "prompt_id": resp.json().get("prompt_id")}
                except Exception as e:
                    result["comfyui"] = {"status": "error", "error": str(e)}
            else:
                result["comfyui"] = {"status": "dry_run", "note": "ComfyUI 未运行，安装后可自动提交"}

            results.append(result)

    return {
        "project": project.name,
        "code": project.code,
        "comfyui_available": comfy_available,
        "comfyui_url": comfy_url,
        "total_renders": len(results),
        "results": results,
    }

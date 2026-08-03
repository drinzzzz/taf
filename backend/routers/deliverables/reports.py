"""
TAF deliverables — 数据报表：BOQ + 优先级矩阵 + 交叉评估
"""
import io, json
from uuid import UUID
from datetime import datetime
from typing import List

from fastapi import APIRouter, Depends, Query, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from services.evaluation import EvaluationEngine
from deps import get_db
from .helpers import _get_project, _get_standard, _get_facilities, _get_all_standards

router = APIRouter(prefix="/api/projects", tags=["策划成果"])


# ═══════════════════════════════════════════
# T1: 设施配置清单 BOQ
# ═══════════════════════════════════════════

def _build_boq_workbook(project, standard, facilities):
    """构建 BOQ workbook（共享：端点+打包器内联均调用）"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    item_index = {i["id"]: i for i in standard.config.get("items", [])}
    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}

    groups = {}
    for f in facilities:
        groups.setdefault(f.category or "??", []).append(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设施配置清单"

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    group_font = Font(name="微软雅黑", bold=True, size=11, color="2F5496")
    group_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:K1")
    ws["A1"] = f"设施配置清单 — {project.name} ({project.code})"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = f"标准: {standard.name} | 生成: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="微软雅黑", size=9, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["序号", "板块", "编号", "设施名称", "类型", "建议品牌/规格", "数量", "参考单价", "小计", "布点区域", "备注"]
    col_widths = [6, 18, 12, 28, 8, 32, 6, 10, 10, 14, 20]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    row = 5; seq = 0; grand_total = 0
    for cat in sorted(groups.keys()):
        cat_name = cat_names.get(cat, cat)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        c = ws.cell(row=row, column=1, value=f"{cat} — {cat_name}")
        c.font = group_font; c.fill = group_fill; c.border = thin_border
        for cc in range(1, 12):
            ws.cell(row=row, column=cc).border = thin_border
            ws.cell(row=row, column=cc).fill = group_fill
        row += 1
        cat_subtotal = 0
        for f in groups[cat]:
            seq += 1
            spec = f.spec or {}
            brands = " / ".join(spec.get("brands", [])) or ""
            spec_text = brands or (json.dumps(spec, ensure_ascii=False) if spec else "")
            price = f.price or 0
            subtotal = price * (f.quantity or 0)
            cat_subtotal += subtotal
            grand_total += subtotal
            pos = f.position or {}
            area = pos.get("space_name", "") or ""
            vals = [seq, cat_name, f.standard_item_id, f.name, f.type or "",
                    spec_text, f.quantity or 0, price if price else "",
                    subtotal if subtotal else "", area, f.notes or ""]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = normal_font; c.border = thin_border
                c.alignment = center_align if ci in (1,5,7,8,9) else left_align
            row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f"{cat_name} 小计")
        c.font = Font(name="微软雅黑", bold=True, size=10); c.border = thin_border
        for cc in range(1, 9):
            ws.cell(row=row, column=cc).border = thin_border
        cs = ws.cell(row=row, column=9, value=cat_subtotal if cat_subtotal else "")
        cs.font = Font(name="微软雅黑", bold=True, size=10); cs.border = thin_border; cs.alignment = center_align
        row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value="总  计")
    c.font = Font(name="微软雅黑", bold=True, size=12, color="C00000"); c.border = thin_border
    for cc in range(1, 9):
        ws.cell(row=row, column=cc).border = thin_border
    cg = ws.cell(row=row, column=9, value=grand_total if grand_total else "")
    cg.font = Font(name="微软雅黑", bold=True, size=12, color="C00000"); cg.border = thin_border; cg.alignment = center_align

    return wb


@router.get("/{project_id}/deliverables/boq")
async def export_boq(
    project_id: UUID,
    format: str = Query("xlsx", regex="^(xlsx|json)$"),
    db: AsyncSession = Depends(get_db),
):
    """导出设施配置清单 (Bill of Quantities)"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    if format == "json":
        item_index = {i["id"]: i for i in standard.config.get("items", [])}
        cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}
        groups = {}
        for f in facilities:
            cat = f.category or "??"
            groups.setdefault(cat, []).append(f)
        result = []
        seq = 0
        for cat in sorted(groups.keys()):
            cat_total = 0
            items = []
            for f in groups[cat]:
                seq += 1
                spec = f.spec or {}
                brands = " / ".join(spec.get("brands", [])) or ""
                price = f.price or 0
                subtotal = price * (f.quantity or 0)
                cat_total += subtotal
                items.append({
                    "seq": seq, "category": cat, "category_name": cat_names.get(cat, cat),
                    "item_id": f.standard_item_id, "name": f.name, "type": f.type,
                    "brands": brands, "spec_detail": spec, "quantity": f.quantity,
                    "unit_price": price, "subtotal": subtotal,
                    "supplier": f.supplier or "", "notes": f.notes or "",
                })
            result.append({"category": cat, "name": cat_names.get(cat, cat), "items": items, "category_subtotal": cat_total})
        return {"project": project.name, "code": project.code, "standard": standard.code, "total_facilities": len(facilities), "groups": result}

    wb = _build_boq_workbook(project, standard, facilities)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    from urllib.parse import quote
    filename = f"{project.code}_设施配置清单.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ═══════════════════════════════════════════
# T2: 改造优先级矩阵
# ═══════════════════════════════════════════

def _build_priority_matrix_core(standard, facilities):
    """共享核心：计算优先级矩阵行数据。端点和打包器共用。"""
    cat_weights = {c["id"]: c["weight"] for c in standard.config.get("categories", [])}
    cat_names = {c["id"]: c["name"] for c in standard.config.get("categories", [])}

    rows = []
    for f in facilities:
        spec = f.spec or {}
        base_impact = 10 if f.type == "prerequisite" else 5
        weight = cat_weights.get(f.category, 0.15)
        impact = round(base_impact * (1 + weight), 1)
        cost_level = spec.get("cost_level", "中")
        cost_score = {"高": 3, "中": 2, "低": 1}.get(cost_level, 2)
        feasibility_map = {"installed": 3, "confirmed": 2, "selected": 1, "draft": 0}
        feasibility = feasibility_map.get(f.status, 0)
        if impact >= 7 and cost_score <= 2 and feasibility >= 2:
            phase = "🔥 近期"
        elif impact >= 5 or (impact >= 7 and cost_score == 3):
            phase = "🟡 中期"
        else:
            phase = "🟢 远期"
        rows.append({
            "name": f.name, "category": f.category,
            "category_name": cat_names.get(f.category, f.category),
            "impact": impact, "cost_level": cost_level, "cost_score": cost_score,
            "feasibility": feasibility, "feasibility_label": f.status,
            "phase": phase,
            "reason": f"{'必选项' if f.type == 'prerequisite' else '加分项'} | 成本{cost_level} | 状态{f.status}",
        })

    phase_order = {"🔥 近期": 0, "🟡 中期": 1, "🟢 远期": 2}
    rows.sort(key=lambda r: (phase_order.get(r["phase"], 9), -r["impact"]))

    summary = {}
    for r in rows:
        summary.setdefault(r["phase"], 0)
        summary[r["phase"]] += 1

    return {"rows": rows, "summary": summary}


def _build_priority_xlsx(project_name, rows):
    """共享：生成优先级矩阵 XLSX BytesIO。端点和打包器共用。"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    phase_fills = {
        "🔥 近期": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "🟡 中期": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
        "🟢 远期": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    }
    normal_font = Font(name="微软雅黑", size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    center_align = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "改造优先级矩阵"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"改造优先级矩阵 — {project_name}"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    summary_parts = []
    for phase in ["🔥 近期", "🟡 中期", "🟢 远期"]:
        count = len([r for r in rows if r["phase"] == phase])
        summary_parts.append(f"{phase}: {count}项")
    ws["A2"] = " | ".join(summary_parts)
    ws["A2"].font = Font(name="微软雅黑", size=10, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["设施名称", "板块", "影响度", "成本", "可行性", "优先级", "阶段建议", "分析理由"]
    widths = [28, 20, 8, 8, 8, 8, 12, 36]
    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row = 5
    for r in rows:
        values = [r["name"], r["category_name"], r["impact"], r["cost_level"],
                  r["feasibility_label"], r["impact"] / 10, r["phase"], r["reason"]]
        fill = phase_fills.get(r["phase"])
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx in (3, 4, 5, 6, 7) else Alignment(horizontal="left", vertical="center")
            if fill:
                cell.fill = fill
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/{project_id}/deliverables/priority-matrix")
async def export_priority_matrix(
    project_id: UUID,
    format: str = Query("xlsx", regex="^(xlsx|json)$"),
    db: AsyncSession = Depends(get_db),
):
    """导出改造优先级矩阵（四象限分析）"""
    project = await _get_project(project_id, db)
    standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    pm = _build_priority_matrix_core(standard, facilities)
    rows = pm["rows"]

    if format == "json":
        return {
            "project": project.name, "code": project.code,
            "summary": pm["summary"],
            "items": rows,
        }

    output = _build_priority_xlsx(project.name, rows)

    from urllib.parse import quote
    filename = f"{project.code}_改造优先级矩阵.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ═══════════════════════════════════════════
# T3: 多标准交叉评估
# ═══════════════════════════════════════════

@router.post("/{project_id}/deliverables/benchmark")
async def benchmark_standards(
    project_id: UUID,
    target_codes: List[str] = Body(..., embed=True),
    db: AsyncSession = Depends(get_db),
):
    """多标准交叉评估：一套设施跑多套标准"""
    project = await _get_project(project_id, db)
    current_standard = await _get_standard(project, db)
    facilities = await _get_facilities(project_id, db)

    all_standards = await _get_all_standards(db)
    std_map = {s.code: s for s in all_standards}

    engine = EvaluationEngine(current_standard.config)
    current_result = engine.calculate_score(
        facilities=[{
            "standard_item_id": f.standard_item_id,
            "type": f.type, "category": f.category,
            "status": f.status, "quantity": f.quantity,
        } for f in facilities],
    )

    results = [{
        "code": current_standard.code, "name": current_standard.name,
        "total_score": current_result["total_score"], "level": current_result["level"],
        "stars": current_result["stars"], "is_current": True,
    }]

    facility_item_ids = {f.standard_item_id for f in facilities}

    for code in target_codes:
        if code == current_standard.code:
            continue
        std = std_map.get(code)
        if not std:
            continue
        engine = EvaluationEngine(std.config)
        result = engine.calculate_score(
            facilities=[{
                "standard_item_id": f.standard_item_id,
                "type": f.type, "category": f.category,
                "status": f.status, "quantity": f.quantity,
            } for f in facilities],
        )
        results.append({
            "code": std.code, "name": std.name,
            "total_score": result["total_score"], "level": result["level"],
            "stars": result["stars"], "is_current": False,
        })

    all_item_ids = set()
    std_items_map = {}
    for std_code in [current_standard.code] + target_codes:
        std = std_map.get(std_code)
        if not std:
            continue
        items = {i["id"]: i for i in std.config.get("items", [])}
        std_items_map[std_code] = items
        all_item_ids.update(items.keys())

    matrix = []
    for item_id in sorted(all_item_ids):
        row_data = {"item_id": item_id}
        covered_count = 0
        for std_code in [current_standard.code] + target_codes:
            items = std_items_map.get(std_code, {})
            in_standard = item_id in items
            has_facility = item_id in facility_item_ids
            row_data[std_code] = "✅" if (in_standard and has_facility) else ("⬜" if in_standard else "—")
            if in_standard and has_facility:
                covered_count += 1
        row_data["covered"] = covered_count
        matrix.append(row_data)

    gap_lists = {}
    for std_code in target_codes:
        if std_code == current_standard.code:
            continue
        items = std_items_map.get(std_code, {})
        gaps = []
        for item_id, item in items.items():
            if item_id not in facility_item_ids and not item.get("is_optional_facility"):
                gaps.append({"item_id": item_id, "name": item["name"], "type": item.get("type", ""), "category": item.get("category", "")})
        if gaps:
            gap_lists[std_code] = gaps

    return {
        "project": project.name, "code": project.code,
        "current_standard": current_standard.code,
        "results": results, "matrix": matrix, "gap_lists": gap_lists,
    }
